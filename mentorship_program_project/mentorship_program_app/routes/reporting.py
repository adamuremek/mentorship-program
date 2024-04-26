"""
/*********************************************************************/
/* FILE NAME: reporting.py                                           */
/*********************************************************************/
/* PART OF PROJECT: Mentorship Program                               */
/*********************************************************************/
/* WRITTEN BY: Andrew, Jordan and David                              */
/* DATE CREATED:                                                     */
/*********************************************************************/
/* PROJECT PURPOSE:                                                  */
/*                                                                   */
/* This project aims to enable comprehensive monitoring and reporting*/
/* within the mentorship platform, facilitating detailed analytics on*/
/* user activities, registrations, and system interactions.          */
/*********************************************************************/
/* FILE PURPOSE:                                                     */
/*                                                                   */
/* This file contains the logic for generating detailed reports on   */
/* various metrics within the system such as user activities, mentor */
/* and mentee registration rates, and overall system usage statistics.*/
/* The data is formatted and outputted as Excel spreadsheets for     */
/* administrative use.                                               */
/*********************************************************************/
/* DJANGO IMPORTED VARIABLE LIST (Alphabetically):                   */
/*                                                                   */
/* - HttpRequest: For obtaining details of HTTP requests             */
/* - HttpResponse: Class for sending HTTP responses                  */
/* - Workbook: For creating and managing Excel workbooks             */
/* - Table, TableStyleInfo: For styling Excel tables within workbooks*/
/*********************************************************************/
/* MODIFICATION HISTORY:                                             */
/*********************************************************************/
/* Date       | Changed By | Changes Made                            */
/* -----------|------------|---------------------------------------- */
/*********************************************************************/
"""


from django.http import HttpRequest, HttpResponse
from datetime import datetime, timedelta

from django.db.models import Count, Value
from django.db.models.functions import Concat
from ..models import Mentee, User
from ..models import SystemLogs
from ..models import UserReport
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.utils import timezone
import os
from io import BytesIO

def get_project_time_statistics():
    """
        Authors Andrew P. Jordan A.
        Collect and synthesize data from Systemlogs and output results based on timespans
    """
    now = timezone.localtime(timezone.now()).date()
    week_ago_date = now - timedelta(days=7)
    month_ago_date = now - timedelta(days=30)

    daily_stats = (
        SystemLogs.objects.filter(str_event=SystemLogs.Event.LOGON_EVENT, cls_log_created_on=now).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_REGISTER_EVENT, cls_log_created_on=now).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_REGISTER_EVENT, cls_log_created_on=now).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT, cls_log_created_on=now).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_DEACTIVATED_EVENT, cls_log_created_on=now).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_DEACTIVATED_EVENT, cls_log_created_on=now).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTORSHIP_TERMINATED_EVENT, cls_log_created_on=now).count()
    )

    weekly_stats = (
        SystemLogs.objects.filter(str_event=SystemLogs.Event.LOGON_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_REGISTER_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_REGISTER_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_DEACTIVATED_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_DEACTIVATED_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTORSHIP_TERMINATED_EVENT, cls_log_created_on__gte=week_ago_date).count()
    )

    monthly_stats = (
        SystemLogs.objects.filter(str_event=SystemLogs.Event.LOGON_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_REGISTER_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_REGISTER_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_DEACTIVATED_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_DEACTIVATED_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTORSHIP_TERMINATED_EVENT, cls_log_created_on__gte=month_ago_date).count()
    )

    lifetime_stats = (
        SystemLogs.objects.filter(str_event=SystemLogs.Event.LOGON_EVENT).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_REGISTER_EVENT).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_REGISTER_EVENT).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_DEACTIVATED_EVENT,).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_DEACTIVATED_EVENT,).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTORSHIP_TERMINATED_EVENT).count()
    )

    return {"Daily":daily_stats, "Weekly":weekly_stats, "Monthly":monthly_stats, "Lifetime":lifetime_stats}

# TODO sorry this look like this. If anyone got free time can you make this "feel" better thanks! -JA
def get_project_overall_statistics():
    """
        Authors Andrew P. Jordan A.
        Collect and synthesize data from Systemlogs based on general stats
    """

    # Query to count the number of Mentors who are currently inactive
    inactive_mentors_count = User.objects.filter(str_role='Mentor', bln_active=False).count()
    # Number of total mentees
    total_mentees = User.objects.filter(str_role='Mentee').count()
    # Number of total mentors
    total_mentors = User.objects.filter(str_role='Mentor').count()
    # Total amount of approved mentorships
    total_approved_mentorships = SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT).count()
    # Total amount of requested mentorships
    total_requested_mentorships = SystemLogs.objects.filter(str_event=SystemLogs.Event.REQUEST_MENTORSHIP_EVENT).count()

    # count number of unassigned mentees
    unassigned_mentees = Mentee.objects.filter(mentor=None).count()
    assigned_mentors = User.objects.all().filter(str_role='Mentor').annotate(mentee_count=Count('mentor___mentee_set')).exclude(mentee_count=0).count()

    return {
        "active_mentees"               : User.objects.filter(str_role='Mentee', bln_active=True).count(),
        "assigned_mentees"             : total_mentees - unassigned_mentees,
        "unassigned_mentees"           : User.objects.filter(str_role='Mentee', bln_active=True).count() - (total_mentees - unassigned_mentees),
        "inactive_mentees"             : User.objects.filter(str_role='Mentee', bln_active=False).count(),
        
        "active_mentors"               :  User.objects.filter(str_role='Mentor', bln_active=True).count(),
        "assigned_mentors"             : assigned_mentors,
        "unassigned_mentors"           : User.objects.filter(str_role='Mentor', bln_active=True).count() - assigned_mentors,
        "inactive_mentors"             : inactive_mentors_count,

        "mentees_per_mentor"           : f"{round(total_mentees/total_mentors,2)}" if total_mentors != 0 else 'N/A',
        "mentor_retention_rate"        : f"{round(100 - ((inactive_mentors_count / total_mentors) * 100))}%" if total_mentors != 0 else 'N/A',
        "successful_match_rate"        : f"{round(total_approved_mentorships / total_requested_mentorships * 100)}%" if total_requested_mentorships != 0 else "N/A",
        
        "pending_mentors"              : User.objects.filter(str_role='MentorPending').count(),
        "unresolved_reports"           : UserReport.objects.filter(bln_resolved = False).count()
    }

def create_systemlog_sheet(workbook: Workbook):
    """
        Authors Jordan A.
        Collects all the systemlogs and creates a sheet supplied with that data
    """
    worksheet = workbook.create_sheet("Report Logs")
    
    # Get all SystemLogs
    # Order them by: Event Name, Date Created On
    # Selecting: Event Name, Date Created On, User ID, User's Full Name
    all_system_logs = SystemLogs.objects.order_by("-cls_log_created_on_sortable").annotate(
    full_name=Concat('specified_user__str_first_name', Value(' '), 'specified_user__str_last_name')
    ).values_list('str_event', 'cls_log_created_on', 'specified_user__id', 'full_name', 'str_details')

    # Append header info
    worksheet.append(["Event", "Log Created On", "User ID", "User Full Name", "Details"])

    # Used to store the widths of the columns, populated by loop below
    # Default values are based on header lengths
    column_widths = [7.57, 16.14, 8.86, 16.14, 5.82]

    # Iterate through all logs, add them to the sheet, then record their text length
    for index, log in enumerate(all_system_logs, 1):
        worksheet.append(log) 

        # Retrieve lengths of the data
        data_widths = [len(str(worksheet[f"A{index}"].value).strip()), 
                      len(str(worksheet[f"B{index}"].value).strip()), 
                      len(str(worksheet[f"C{index}"].value).strip()),
                      len(str(worksheet[f"D{index}"].value).strip()),
                      len(str(worksheet[f"E{index}"].value).strip())]
        
        # Checking if the new widths surpasses the previous width, if so, replace it
        column_widths[0] = data_widths[0] if data_widths[0] > column_widths[0] else column_widths[0]
        column_widths[1] = data_widths[1] if data_widths[1] > column_widths[1] else column_widths[1]
        column_widths[2] = data_widths[2] if data_widths[2] > column_widths[2] else column_widths[2]
        column_widths[3] = data_widths[3] if data_widths[3] > column_widths[3] else column_widths[3]
        column_widths[4] = data_widths[4] if data_widths[4] > column_widths[4] else column_widths[4]

    # Resize columns to fit the data 
    worksheet.column_dimensions["A"].width = column_widths[0] + 2
    worksheet.column_dimensions["B"].width = column_widths[1] + 2
    worksheet.column_dimensions["C"].width = column_widths[2] + 2
    worksheet.column_dimensions["D"].width = column_widths[3] + 2
    worksheet.column_dimensions["E"].width = column_widths[4] + 2

    # Apply autofilter to the columns
    table_range = f"A1:E{len(all_system_logs) + 1}"
    
    table = Table(displayName="SystemLogs", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    worksheet.add_table(table)

def create_mentorship_statistics_sheet(workbook: Workbook):
    """
        Authors Jordan A.
        Collects general stats and creates a sheet supplied with that data
    """
    worksheet = workbook.create_sheet("Mentorship Statistics")
    
    timespans = get_project_time_statistics()
    overall_stats = get_project_overall_statistics()

    header = ["Daily Stats", "Weekly Stats", "Monthly Stats", "Lifetime stats", "Program stats"]
    worksheet.merge_cells('A1:B1')
    worksheet.merge_cells('D1:E1')
    worksheet.merge_cells('G1:H1')
    worksheet.merge_cells('J1:K1')
    worksheet.merge_cells('A9:B9')
    
    worksheet["A1"] = header[0]
    worksheet["D1"] = header[1]
    worksheet["G1"] = header[2]
    worksheet["J1"] = header[3]
    worksheet["A9"] = header[4]

    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet["D1"].alignment = Alignment(horizontal="center")
    worksheet["G1"].alignment = Alignment(horizontal="center")
    worksheet["J1"].alignment = Alignment(horizontal="center")
    worksheet["A9"].alignment = Alignment(horizontal="center")

    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['G'].width = 40
    worksheet.column_dimensions['J'].width = 40

    for column_1, column_2, timespan in (("A", "B", "Daily"), ("D", "E", "Weekly"), ("G", "H", "Monthly"), ("J", "K", "Lifetime")):
        
        worksheet[f"{column_1}2"] = f"{timespan} Visitors"
        worksheet[f"{column_1}3"] = f"{timespan} Mentee Signup"
        worksheet[f"{column_1}4"] = f"{timespan} Mentor Signup"
        worksheet[f"{column_1}5"] = f"{timespan} Assigned Mentees"
        worksheet[f"{column_1}6"] = f"{timespan} Deactivated Mentees"
        worksheet[f"{column_1}7"] = f"{timespan} Deactivated Mentors"
        
        worksheet[f"{column_2}2"] = timespans[timespan][0]
        worksheet[f"{column_2}3"] = timespans[timespan][1]
        worksheet[f"{column_2}4"] = timespans[timespan][2]
        worksheet[f"{column_2}5"] = timespans[timespan][3]
        worksheet[f"{column_2}6"] = timespans[timespan][4]
        worksheet[f"{column_2}7"] = timespans[timespan][5]

    for index, stat in enumerate(overall_stats, start=10):
        worksheet[f"B{index}"].alignment = Alignment(horizontal="right")
        worksheet[f"A{index}"] = stat.replace("_", " ").title()
        worksheet[f"B{index}"] = overall_stats[stat]

def generate_report(req : HttpRequest):
    """
    Authors Andrew P. Jordan A.
    Generates a report (duh!)
    """
    workbook = Workbook()
    workbook.remove(workbook.active) 
    create_mentorship_statistics_sheet(workbook)
    create_systemlog_sheet(workbook)
    file_name = f"Mentorship-Program-statistics-{datetime.now().strftime('%Y.%m.%d.%H.%M.%S')}.xlsx"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',   
    )
    
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response


