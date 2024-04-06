from django.http import HttpRequest, FileResponse, HttpResponse
from django.conf import settings
from datetime import date, datetime, timedelta

from django.db.models import Count, Q, Value,Case,BooleanField,When,F,OuterRef,Subquery
from ..models import Mentee, MentorshipRequest, User
from ..models import Mentor
from ..models import SystemLogs
from ..models import UserReport
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.utils import timezone
import os
from io import BytesIO

def get_project_time_statistics():
    """
        Authors Andrew P. Jordan A.
        Collect and synthesize data from Systemlogs and output results based on timespans
    """
    week_ago_date = timezone.now().date() - timedelta(days=7)
    month_ago_date = timezone.now().date() - timedelta(days=30)


    daily_stats = (
        SystemLogs.objects.filter(str_event=SystemLogs.Event.LOGON_EVENT, cls_log_created_on=timezone.now().date()).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_REGISTER_EVENT, cls_log_created_on=timezone.now().date()).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_REGISTER_EVENT, cls_log_created_on=timezone.now().date()).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT, cls_log_created_on=timezone.now().date()).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_DEACTIVATED_EVENT, cls_log_created_on=timezone.now().date()).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_DEACTIVATED_EVENT, cls_log_created_on=timezone.now().date()).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTORSHIP_TERMINATED_EVENT, cls_log_created_on=timezone.now().date()).count()
    )

    weekly_stats = (
        SystemLogs.objects.filter(str_event=SystemLogs.Event.LOGON_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_REGISTER_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_REGISTER_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_DEACTIVATED_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_DEACTIVATED_EVENT, cls_log_created_on__gte=week_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTORSHIP_TERMINATED_EVENT, cls_log_created_on=week_ago_date).count()
    )

    monthly_stats = (
        SystemLogs.objects.filter(str_event=SystemLogs.Event.LOGON_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_REGISTER_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_REGISTER_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_DEACTIVATED_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_DEACTIVATED_EVENT, cls_log_created_on__gte=month_ago_date).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTORSHIP_TERMINATED_EVENT, cls_log_created_on=month_ago_date).count()
    )

    lifetime_stats = (
        SystemLogs.objects.filter(str_event=SystemLogs.Event.LOGON_EVENT).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_REGISTER_EVENT).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_REGISTER_EVENT).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTEE_DEACTIVATED_EVENT,).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTOR_DEACTIVATED_EVENT,).count(),
        SystemLogs.objects.filter(str_event=SystemLogs.Event.MENTORSHIP_TERMINATED_EVENT).count()
    )

    return {"Daily":daily_stats, "Weekly":weekly_stats, "Monthly":monthly_stats, "Lifetime":lifetime_stats}

# TODO sorry this look like this. If anyone got free time can you make this "feel" better thanks! -JA
def get_project_overall_statistics():
    """
        Authors Andrew P. Jordan A.
        Collect and synthesize data from Systemlogs based on general stats
    """

    inactive_date = datetime.now() - timedelta(days=365)
    # Query to count the number of Mentors who are currently inactive
    inactive_mentors_count = User.objects.filter(str_role='Mentor', bln_active=False).count()
    # Query to count the number of Mentors who are currently active
    active_mentors_count = User.objects.filter(str_role='Mentor', bln_active=True).count()
    # Number of total mentees
    total_mentees = User.objects.filter(str_role='Mentee').count()
    # Number of total mentors
    total_mentors = User.objects.filter(str_role='Mentor').count()
    # Total amount of approved mentorships
    total_approved_mentorships = SystemLogs.objects.filter(str_event=SystemLogs.Event.APPROVE_MENTORSHIP_EVENT).count()
    # Total amount of requested mentorships
    total_requested_mentorships = SystemLogs.objects.filter(str_event=SystemLogs.Event.REQUEST_MENTORSHIP_EVENT).count()


    # count number of unassigned mentees
    unassigned_mentees = Mentee.objects.filter(mentor=None).count()
    assigned_mentors = User.objects.all().filter(str_role='Mentor').annotate(mentee_count=Count('mentor___mentee_set')).exclude(mentee_count=0).count()

    return {
        "active_mentees"               : User.objects.filter(str_role='Mentee', bln_active=True).count(),
        "assigned_mentees"             : total_mentees - unassigned_mentees,
        "unassigned_mentees"           : unassigned_mentees,
        "inactive_mentees"             : User.objects.filter(str_role='Mentee', bln_active=False).count(),
        
        "active_mentors"               :  User.objects.filter(str_role='Mentor', bln_active=True).count(),
        "assigned_mentors"             : assigned_mentors,
        "unassigned_mentors"           : total_mentors - assigned_mentors,
        "inactive_mentors"             : inactive_mentors_count,

        "mentees_per_mentor"           : f"{round(total_mentees/total_mentors,2)}" if total_mentors != 0 else 'N/A',
        "mentor_retention_rate"        : f"{round(100 - ((inactive_mentors_count / total_mentors) * 100))}%" if total_mentors != 0 else 'N/A',
        "successful_match_rate"        : f"{round(total_approved_mentorships / total_requested_mentorships * 100)}%" if total_requested_mentorships != 0 else "N/A",
        
        "pending_mentors"              : User.objects.filter(str_role='MentorPending').count(),
        "unresolved_reports"           : UserReport.objects.filter(bln_resolved = False).count()
    }

def generate_report(req : HttpRequest):
    """
    Authors Andrew P. Jordan A.
    Generate report (duh!)
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Mentorship Statistics"
    timespans = get_project_time_statistics()
    overall_stats = get_project_overall_statistics()

    header = ["Daily Stats", "Weekly Stats", "Monthly Stats", "Lifetime stats", "Program stats"]
    worksheet.merge_cells('A1:B1')
    worksheet.merge_cells('D1:E1')
    worksheet.merge_cells('G1:H1')
    worksheet.merge_cells('J1:K1')
    worksheet.merge_cells('A9:B9')
    
    worksheet["A1"] = header[0]
    worksheet["D1"] = header[1]
    worksheet["G1"] = header[2]
    worksheet["J1"] = header[3]
    worksheet["A9"] = header[4]

    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet["D1"].alignment = Alignment(horizontal="center")
    worksheet["G1"].alignment = Alignment(horizontal="center")
    worksheet["J1"].alignment = Alignment(horizontal="center")
    worksheet["A9"].alignment = Alignment(horizontal="center")

    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['G'].width = 40
    worksheet.column_dimensions['J'].width = 40

    for column_1, column_2, timespan in (("A", "B", "Daily"), ("D", "E", "Weekly"), ("G", "H", "Monthly"), ("J", "K", "Lifetime")):
        
        worksheet[f"{column_1}2"] = f"{timespan} Visitors"
        worksheet[f"{column_1}3"] = f"{timespan} Mentee Signup"
        worksheet[f"{column_1}4"] = f"{timespan} Mentor Signup"
        worksheet[f"{column_1}5"] = f"{timespan} Assigned Mentees"
        worksheet[f"{column_1}6"] = f"{timespan} Deactivated Mentees"
        worksheet[f"{column_1}7"] = f"{timespan} Deactivated Mentors"
        
        worksheet[f"{column_2}2"] = timespans[timespan][0]
        worksheet[f"{column_2}3"] = timespans[timespan][1]
        worksheet[f"{column_2}4"] = timespans[timespan][2]
        worksheet[f"{column_2}5"] = timespans[timespan][3]
        worksheet[f"{column_2}6"] = timespans[timespan][4]
        worksheet[f"{column_2}7"] = timespans[timespan][5]

    for index, stat in enumerate(overall_stats, start=10):
        worksheet[f"B{index}"].alignment = Alignment(horizontal="right")
        worksheet[f"A{index}"] = stat.replace("_", " ").title()
        worksheet[f"B{index}"] = overall_stats[stat]

    file_name = f"Mentorship-Program-statistics-{datetime.now().strftime('%Y.%m.%d.%H.%M.%S')}.xlsx"

    report_logs = workbook.create_sheet("Report Logs")
    report


    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',   
    )
    
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    # response['Refresh'] = '0; url=' + req.path
    return response


